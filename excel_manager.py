import os
import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import EXCEL_FILE_PATH, CACHE_FILE_PATH

# ألوان وتنسيقات الهيئة العامة للرعاية الصحية
NAVY_HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
TEAL_HEADER_FILL = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
GOLD_HEADER_FILL = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")
RED_HEADER_FILL  = PatternFill(start_color="A82020", end_color="A82020", fill_type="solid")
GREEN_HEADER_FILL= PatternFill(start_color="237804", end_color="237804", fill_type="solid")

ZEBRA_EVEN_FILL  = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
ZEBRA_ODD_FILL   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

STATUS_UNREPLIED_FILL = PatternFill(start_color="FFEAE6", end_color="FFEAE6", fill_type="solid")
STATUS_REPLIED_FILL   = PatternFill(start_color="E6FFFB", end_color="E6FFFB", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REGULAR_FONT = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", size=10, bold=True)
LINK_FONT    = Font(name="Calibri", size=10, underline="single", color="0563C1")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def save_cache_data(data: dict):
    """حفظ البيانات في كاش محلي للوصول فائق السرعة عبر الويب"""
    try:
        with open(CACHE_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving cache: {e}")

def load_cache_data() -> dict:
    """تحميل البيانات من الكاش المحلي"""
    if CACHE_FILE_PATH.exists():
        try:
            with open(CACHE_FILE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading cache: {e}")
    return {"inbox": [], "sent": [], "attachments": [], "last_sync": None}

def format_sheet_common(ws, headers: list, header_fill=NAVY_HEADER_FILL):
    """تنسيق مشترك لأوراق العمل باللغة العربية مع RTL"""
    ws.sheet_view.rightToLeft = True
    
    # كتابة الترويسات
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=h_text)
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        
    ws.row_dimensions[1].height = 28

def auto_fit_columns(ws, max_widths=None):
    """ضبط عرض الأعمدة تلقائياً ليناسب المحتوى بدون قص"""
    if max_widths is None:
        max_widths = {}
        
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            # حساب تقريبي لطول الحروف العربية
            cell_len = len(val_str.encode('utf-8')) // 2 if any(ord(c) > 127 for c in val_str) else len(val_str)
            if cell_len > max_len:
                max_len = cell_len
                
        # حد أقصى وحد أدنى للعرض
        cap = max_widths.get(col_letter, 55)
        calculated_width = max(max_len + 3, 12)
        ws.column_dimensions[col_letter].width = min(calculated_width, cap)

def write_excel_registry(inbox_emails: list, sent_emails: list, attachments: list) -> str:
    """
    إنشاء وتحديث ملف الإكسيل الاحترافي الكامل:
    - ورقة 1: 📅 تقرير الحصر بالتواريخ (الوارد والصادر المعلق)
    - ورقة 2: 📥 وارد - لم يتم الرد عليه
    - ورقة 3: 📤 صادر - بانتظار رد الجهات
    - ورقة 4: ✅ وارد - تم الرد عليه
    - ورقة 5: 📨 صادر - تم استلام رده
    - ورقة 6: 📋 كافة الرسائل (وارد وصادر)
    - ورقة 7: 📎 سجل المرفقات (مع روابط قابلة للنقر)
    - ورقة 8: 📊 تقرير الحصر والمؤشرات
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    inbox_unreplied = [e for e in inbox_emails if e.get("status_code") == "inbox_unreplied"]
    inbox_replied = [e for e in inbox_emails if e.get("status_code") == "inbox_replied"]
    
    sent_unreplied = [e for e in sent_emails if e.get("status_code") == "sent_unreplied"]
    sent_replied = [e for e in sent_emails if e.get("status_code") == "sent_replied"]
    
    # -------------------------------------------------------------
    # 1. ورقة: 📅 تقرير الحصر بالتواريخ (الوارد والصادر المعلق معاً)
    # -------------------------------------------------------------
    ws_chrono = wb.create_sheet(title="📅 تقرير المعلق بالتواريخ")
    headers_chrono = [
        "م", "التاريخ والوقت", "النوع", "الطرف الآخر (المرسل/المستلم)", "الموضوع", 
        "الملخص والتكليف المطلوب", "المرفقات", "الأيام المنقضية", "الموقف والمتابعة"
    ]
    format_sheet_common(ws_chrono, headers_chrono, header_fill=NAVY_HEADER_FILL)
    
    # دمج الوارد والصادر المعلق وترتيبهما زمنياً
    all_unreplied = []
    for i in inbox_unreplied:
        all_unreplied.append({
            **i,
            "type_label": "📥 وارد لم يُرد عليه",
            "sort_dt": i.get("datetime_received") or "",
            "party": i.get("sender_name") or i.get("sender_email") or "غير معروف",
            "status_note": "معلق لدى المكتب الفني (يتطلب رداً)"
        })
    for s in sent_unreplied:
        recips = ", ".join(s.get("to_recipients_names", []) or s.get("to_recipients_emails", [])) or "الجهات المعنية"
        all_unreplied.append({
            **s,
            "type_label": "📤 صادر بانتظار رد",
            "sort_dt": s.get("datetime_sent") or "",
            "party": recips,
            "status_note": "معلق لدى الجهات الخارجية (لم يرد رد)"
        })
    all_unreplied.sort(key=lambda x: x.get("sort_dt", ""), reverse=True)

    for row_idx, item in enumerate(all_unreplied, 2):
        is_inbox = "وارد" in item.get("type_label", "")
        fill = PatternFill(start_color="FFF1F0" if is_inbox else "FEF9E7", end_color="FFF1F0" if is_inbox else "FEF9E7", fill_type="solid")
        
        days_ago = ""
        dt_str = item.get("sort_dt")
        if dt_str:
            try:
                dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                days_ago = (datetime.now(dt.tzinfo) - dt).days
            except Exception: pass

        summary_text = item.get("summary", "")
        if item.get("actions") and len(item.get("actions")) > 0:
            summary_text = f"📌 {item.get('actions')[0]} \n{summary_text}"

        row_vals = [
            row_idx - 1,
            item.get("sort_dt", "")[:16].replace("T", " "),
            item.get("type_label", ""),
            item.get("party", ""),
            item.get("subject", ""),
            summary_text,
            len(item.get("attachments", [])),
            days_ago,
            item.get("status_note", "")
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_chrono.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = BOLD_FONT if col_idx in [2, 3] else REGULAR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx in [4, 5, 6, 9] else "center", vertical="center", wrap_text=True)

    auto_fit_columns(ws_chrono, {"B": 18, "C": 20, "D": 30, "E": 35, "F": 45, "I": 30})
    
    # -------------------------------------------------------------
    # 1. ورقة: وارد - لم يتم الرد عليه (Inbox Unreplied)
    # -------------------------------------------------------------
    ws_in_un = wb.create_sheet(title="📥 وارد - لم يتم الرد عليه")
    headers_in_un = [
        "م", "تاريخ الاستلام", "اسم المرسل", "بريد المرسل", "الموضوع", 
        "الملخص التنفيذي", "التكليفات المطلوبة", "عدد المرفقات", "الأيام المنقضية"
    ]
    format_sheet_common(ws_in_un, headers_in_un, header_fill=RED_HEADER_FILL)
    
    for row_idx, item in enumerate(inbox_unreplied, 2):
        fill = ZEBRA_EVEN_FILL if row_idx % 2 == 0 else ZEBRA_ODD_FILL
        actions_text = " \n- ".join(item.get("actions", []))
        if actions_text: actions_text = "- " + actions_text
            
        days_ago = ""
        dt_str = item.get("datetime_received")
        if dt_str:
            try:
                dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                days_ago = (datetime.now(dt.tzinfo) - dt).days
            except Exception: pass
                
        row_vals = [
            row_idx - 1,
            item.get("datetime_received", ""),
            item.get("sender_name", ""),
            item.get("sender_email", ""),
            item.get("subject", ""),
            item.get("summary", ""),
            actions_text,
            len(item.get("attachments", [])),
            days_ago
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_in_un.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx in [5, 6, 7] else "center", vertical="center", wrap_text=True)
                                       
    auto_fit_columns(ws_in_un, {"E": 35, "F": 45, "G": 40})

    # -------------------------------------------------------------
    # 2. ورقة: صادر - بانتظار رد الجهات (Sent Unreplied)
    # -------------------------------------------------------------
    ws_sent_un = wb.create_sheet(title="📤 صادر - بانتظار رد الجهات")
    headers_sent_un = [
        "م", "تاريخ الإرسال", "المرسل إليهم (الجهات)", "الموضوع الصادر", "الملخص والمطلوب", "عدد المرفقات", "الأيام منذ الإرسال"
    ]
    format_sheet_common(ws_sent_un, headers_sent_un, header_fill=GOLD_HEADER_FILL)
    
    for row_idx, item in enumerate(sent_unreplied, 2):
        fill = ZEBRA_EVEN_FILL if row_idx % 2 == 0 else ZEBRA_ODD_FILL
        recips = ", ".join(item.get("to_recipients_names", []) or item.get("to_recipients_emails", []))
        
        days_ago = ""
        dt_str = item.get("datetime_sent")
        if dt_str:
            try:
                dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
                days_ago = (datetime.now(dt.tzinfo) - dt).days
            except Exception: pass

        row_vals = [
            row_idx - 1,
            item.get("datetime_sent", ""),
            recips,
            item.get("subject", ""),
            item.get("summary", ""),
            len(item.get("attachments", [])),
            days_ago
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_sent_un.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx in [3, 4, 5] else "center", vertical="center", wrap_text=True)
                                       
    auto_fit_columns(ws_sent_un, {"C": 35, "D": 35, "E": 45})

    # -------------------------------------------------------------
    # 3. ورقة: وارد - تم الرد عليه من المكتب الفني (Inbox Replied)
    # -------------------------------------------------------------
    ws_in_rep = wb.create_sheet(title="✅ وارد - تم الرد عليه")
    headers_in_rep = [
        "م", "تاريخ الاستلام", "اسم المرسل", "الموضوع الأصلي", "الملخص",
        "تاريخ الرد الصادر", "موضوع الرد", "من قام بالرد"
    ]
    format_sheet_common(ws_in_rep, headers_in_rep, header_fill=GREEN_HEADER_FILL)
    
    for row_idx, item in enumerate(inbox_replied, 2):
        fill = ZEBRA_EVEN_FILL if row_idx % 2 == 0 else ZEBRA_ODD_FILL
        row_vals = [
            row_idx - 1,
            item.get("datetime_received", ""),
            item.get("sender_name", "") or item.get("sender_email", ""),
            item.get("subject", ""),
            item.get("summary", ""),
            item.get("reply_datetime", ""),
            item.get("reply_subject", ""),
            item.get("replied_by", "")
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_in_rep.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx in [4, 5, 7] else "center", vertical="center", wrap_text=True)
                                       
    auto_fit_columns(ws_in_rep, {"D": 35, "E": 45, "G": 35})

    # -------------------------------------------------------------
    # 4. ورقة: صادر - تم استلام الرد (Sent Replied)
    # -------------------------------------------------------------
    ws_sent_rep = wb.create_sheet(title="📨 صادر - تم استلام رده")
    headers_sent_rep = [
        "م", "تاريخ الإرسال", "الجهة المرسل إليها", "الموضوع الصادر", "الجهة التي قامت بالرد", "تاريخ استلام الرد", "موضوع رد الجهة"
    ]
    format_sheet_common(ws_sent_rep, headers_sent_rep, header_fill=TEAL_HEADER_FILL)
    
    for row_idx, item in enumerate(sent_replied, 2):
        fill = ZEBRA_EVEN_FILL if row_idx % 2 == 0 else ZEBRA_ODD_FILL
        recips = ", ".join(item.get("to_recipients_names", []) or item.get("to_recipients_emails", []))
        row_vals = [
            row_idx - 1,
            item.get("datetime_sent", ""),
            recips,
            item.get("subject", ""),
            item.get("response_from", ""),
            item.get("response_date", ""),
            item.get("response_subject", "")
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_sent_rep.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx in [3, 4, 7] else "center", vertical="center", wrap_text=True)
                                       
    auto_fit_columns(ws_sent_rep, {"C": 35, "D": 35, "E": 25, "G": 35})

    # -------------------------------------------------------------
    # 5. ورقة: كافة الوارد والصادر (All Emails)
    # -------------------------------------------------------------
    ws_all = wb.create_sheet(title="📋 كافة الرسائل")
    headers_all = [
        "م", "النوع", "الحالة", "التاريخ", "الطرف الآخر", "الموضوع", "الملخص", "المرفقات"
    ]
    format_sheet_common(ws_all, headers_all, header_fill=NAVY_HEADER_FILL)
    
    combined = []
    for i in inbox_emails:
        combined.append({**i, "type_label": "وارد (Inbox)", "sort_date": i.get("datetime_received") or ""})
    for s in sent_emails:
        combined.append({**s, "type_label": "صادر (Sent)", "sort_date": s.get("datetime_sent") or ""})
    combined.sort(key=lambda x: x.get("sort_date", ""), reverse=True)

    for row_idx, item in enumerate(combined, 2):
        fill = ZEBRA_EVEN_FILL if row_idx % 2 == 0 else ZEBRA_ODD_FILL
        party = item.get("sender_name") or item.get("sender_email") if item.get("folder") == "Inbox" else (", ".join(item.get("to_recipients_names", []) or item.get("to_recipients_emails", [])))
        st = item.get("status", "غير محدد")
        
        row_vals = [
            row_idx - 1,
            item.get("type_label", ""),
            st,
            item.get("sort_date", ""),
            party,
            item.get("subject", ""),
            item.get("summary", ""),
            len(item.get("attachments", []))
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = BOLD_FONT if col_idx in [2, 3] else REGULAR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx in [5, 6, 7] else "center", vertical="center", wrap_text=True)
                                       
    auto_fit_columns(ws_all, {"E": 30, "F": 35, "G": 45})

    # -------------------------------------------------------------
    # 6. ورقة: سجل المرفقات (Attachments)
    # -------------------------------------------------------------
    ws_att = wb.create_sheet(title="📎 سجل المرفقات")
    headers_att = [
        "م", "اسم الملف", "نوع الملف", "الحجم (KB)", "موضوع الإيميل", "تاريخ الإيميل", "المرسل", "فتح الملف مباشرة"
    ]
    format_sheet_common(ws_att, headers_att, header_fill=GOLD_HEADER_FILL)
    
    for row_idx, att in enumerate(attachments, 2):
        fill = ZEBRA_EVEN_FILL if row_idx % 2 == 0 else ZEBRA_ODD_FILL
        local_path = att.get("file_path", "")
        
        row_vals = [
            row_idx - 1,
            att.get("name", ""),
            att.get("extension", ""),
            round(att.get("size_bytes", 0) / 1024, 1),
            att.get("email_subject", ""),
            att.get("email_date", ""),
            att.get("email_sender", ""),
            "📁 انقر هنا لفتح الملف" if local_path else "غير متوفر"
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_att.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col_idx == 5 else "center", vertical="center")
            
            if col_idx == 8 and local_path and os.path.exists(local_path):
                cell.hyperlink = local_path
                cell.font = LINK_FONT
            else:
                cell.font = REGULAR_FONT
                
    auto_fit_columns(ws_att, {"B": 30, "E": 40, "G": 25, "H": 25})

    # -------------------------------------------------------------
    # 7. ورقة: لوحة المؤشرات والإحصائيات (KPIs)
    # -------------------------------------------------------------
    ws_kpi = wb.create_sheet(title="📊 لوحة المؤشرات")
    ws_kpi.sheet_view.rightToLeft = True
    
    ws_kpi.merge_cells("A1:C1")
    title_cell = ws_kpi.cell(row=1, column=1, value="تقرير حصر ومتابعة بريد المكتب الفني بالأقصر - الهيئة العامة للرعاية الصحية")
    title_cell.fill = NAVY_HEADER_FILL
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_kpi.row_dimensions[1].height = 35

    total_inbox = len(inbox_emails)
    total_sent = len(sent_emails)
    total_in_replied = len(inbox_replied)
    total_in_unreplied = len(inbox_unreplied)
    total_sent_unreplied = len(sent_unreplied)
    total_sent_replied = len(sent_replied)
    total_attachments = len(attachments)
    
    inbox_reply_rate = round((total_in_replied / total_inbox * 100), 1) if total_inbox > 0 else 0
    sent_reply_rate = round((total_sent_replied / total_sent * 100), 1) if total_sent > 0 else 0
    
    stats_data = [
        ("📥 إجمالي الرسائل الواردة", total_inbox, "كافة الوارد المحصور"),
        ("⏳ الوارد المعلق (لم يتم الرد عليه من المكتب الفني)", total_in_unreplied, "يتطلب اتخاذ إجراء ورد عاجل"),
        ("✅ الوارد المكتمل (تم الرد عليه من المكتب الفني)", total_in_replied, f"نسبة الإنجاز: {inbox_reply_rate}%"),
        ("📤 إجمالي الرسائل الصادرة", total_sent, "مراسلات المكتب الفني للجهات"),
        ("⌛ الصادر المعلق (بانتظار رد الجهات الخارجية/المستشفيات)", total_sent_unreplied, "مراسلات صادرة لم يرد عليها أحد"),
        ("📨 الصادر المردود عليه من الجهات", total_sent_replied, f"نسبة استجابة الجهات: {sent_reply_rate}%"),
        ("📎 إجمالي المرفقات الموثقة والمنزلة", total_attachments, "ملفات PDF، مستندات، صور"),
        ("🕒 تاريخ ووقت التحديث الأخير", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "تحديث تلقائي")
    ]
    
    for r_idx, (label, val, note) in enumerate(stats_data, 3):
        lbl_cell = ws_kpi.cell(row=r_idx, column=1, value=label)
        lbl_cell.font = BOLD_FONT
        lbl_cell.fill = ZEBRA_EVEN_FILL
        lbl_cell.border = THIN_BORDER
        lbl_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        val_cell = ws_kpi.cell(row=r_idx, column=2, value=val)
        val_cell.font = Font(name="Calibri", size=12, bold=True, color="1B365D")
        val_cell.fill = ZEBRA_ODD_FILL
        val_cell.border = THIN_BORDER
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        note_cell = ws_kpi.cell(row=r_idx, column=3, value=note)
        note_cell.font = REGULAR_FONT
        note_cell.fill = ZEBRA_EVEN_FILL
        note_cell.border = THIN_BORDER
        note_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws_kpi.row_dimensions[r_idx].height = 24
        
    ws_kpi.column_dimensions["A"].width = 50
    ws_kpi.column_dimensions["B"].width = 20
    ws_kpi.column_dimensions["C"].width = 35

    # حفظ ملف الإكسيل
    wb.save(EXCEL_FILE_PATH)
    
    # حفظ الكاش
    save_cache_data({
        "inbox": inbox_emails,
        "sent": sent_emails,
        "attachments": attachments,
        "last_sync": datetime.now().isoformat(),
        "stats": {
            "total_inbox": total_inbox,
            "total_sent": total_sent,
            "inbox_replied_count": total_in_replied,
            "inbox_unreplied_count": total_in_unreplied,
            "sent_replied_count": total_sent_replied,
            "sent_unreplied_count": total_sent_unreplied,
            "inbox_reply_rate": inbox_reply_rate,
            "sent_reply_rate": sent_reply_rate,
            "total_attachments": total_attachments
        }
    })
    
    return str(EXCEL_FILE_PATH)
